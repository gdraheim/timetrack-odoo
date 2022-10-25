#! /usr/bin/python3
import logging
from typing import Optional, Union, Dict, List, Any, Sequence
from tabtotext import JSONList, JSONDict, tabToGFM, strNone, strDateTime

from openpyxl import Workbook, load_workbook  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
from openpyxl.styles.cell_style import CellStyle as Style  # type: ignore
from openpyxl.styles.alignment import Alignment  # type: ignore
from openpyxl.styles.numbers import NumberFormat, builtin_format_id  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

import datetime
DayOrTime = (datetime.date, datetime.datetime)

MINWIDTH = 4
MAXCOL = 1000
MAXROWS = 100000

# Excel encodes empty-string as nonexistant cell.
# Since we want to encode None as empty cell (to allow numeric computations), we assign a value for empty-string.
# Note that other people hat recommended "NA()" for null which shows as "N#A" and it is correctly transferred
# into a database as NULL by Excel itself. However in your formulas you would need to skip those cells any
# numeric operation with a numeric value and some NA() returns NA()

_Empty_String = " "

logg = logging.getLogger("TABTOXLSX")

def set_cell(ws: Worksheet, row: int, col: int, value: Any, style: Style) -> None:  # type: ignore
    coordinate = {"column": col + 1, "row": row + 1}
    ws.cell(**coordinate).value = value
    # ws.cell(**coordinate).font = style.font
    # ws.cell(**coordinate).fill = style.fill
    # ws.cell(**coordinate).border = style.border
    ws.cell(**coordinate).alignment = style.alignment
    ws.cell(**coordinate).number_format = style.number_format
    ws.cell(**coordinate).protection = style.protection
def set_width(ws: Worksheet, col: int, width: int) -> None:  # type: ignore
    ws.column_dimensions[get_column_letter(col + 1)].width = width


def saveToXLSXx(filename: str, result: Union[JSONList, JSONDict], sorts: Sequence[str] = [], formats: Dict[str, str] = {}, legend: Union[Dict[str, str], Sequence[str]] = []) -> None:
    if isinstance(result, Dict):
        result = [result]
    saveToXLSX(filename, result, sorts, formats, legend)

def saveToXLSX(filename: str, result: JSONList, sorts: Sequence[str] = [], formats: Dict[str, str] = {}, legend: Union[Dict[str, str], Sequence[str]] = []) -> None:
    def sortkey(header: str) -> str:
        if header in sorts:
            return "%07i" % sorts.index(header)
        return header
    def sortrow(item: JSONDict) -> str:
        sortvalue = ""
        for sort in sorts:
            if sort in item:
                value = item[sort]
                if isinstance(value, int):
                    sortvalue += "\n%020i" % value
                else:
                    sortvalue += "\n" + strDateTime(value)
            else:
                sortvalue += "\n-"
        return sortvalue
    cols: Dict[str, int] = {}
    for item in result:
        for name, value in item.items():
            paren = 0
            if name not in cols:
                cols[name] = max(MINWIDTH, len(name))
            cols[name] = max(cols[name], len(strNone(value)))
    #
    row = 0
    workbook = Workbook()
    ws = workbook.active
    ws.title = "data"
    style = Style()
    txt_style = Style()
    txt_style.number_format = 'General'
    txt_style.alignment = Alignment(horizontal='left')
    dat_style = Style()
    dat_style.number_format = 'd.mm.yy'
    dat_style.alignment = Alignment(horizontal='right')
    num_style = Style()
    num_style.number_format = '#,##0.00'
    num_style.alignment = Alignment(horizontal='right')
    int_style = Style()
    int_style.number_format = '#,##0'
    int_style.alignment = Alignment(horizontal='right')
    col = 0
    for name in sorted(cols.keys(), key=sortkey):
        set_cell(ws, row, col, name, txt_style)
        set_width(ws, col, cols[name] + 1 + int(cols[name]/3))
        col += 1
    row += 1
    for item in sorted(result, key=sortrow):
        values: JSONDict = dict([(name, "") for name in cols.keys()])
        for name, value in item.items():
            values[name] = value
        col = 0
        for name in sorted(cols.keys(), key=sortkey):
            value = values[name]
            if value is None:
                pass
            elif isinstance(value, DayOrTime):
                set_cell(ws, row, col, value, dat_style)
            elif isinstance(value, int):
                set_cell(ws, row, col, value, int_style)
            elif isinstance(value, float):
                set_cell(ws, row, col, value, num_style)
            else:
                if not value:
                    set_cell(ws, row, col, _Empty_String, txt_style)
                else:
                    set_cell(ws, row, col, value, txt_style)
            col += 1
        row += 1
    if legend:
        ws = workbook.create_sheet()
        ws.title = "legend"
        if isinstance(legend, str):
            set_cell(ws, 0, 1, legend, txt_style)
        elif isinstance(legend, dict):
            for row, name in enumerate(sorted(legend.keys(), key=sortkey)):
                set_cell(ws, row, 0, name, txt_style)
                set_cell(ws, row, 1, legend[name], txt_style)
        else:
            for row, line in enumerate(legend):
                set_cell(ws, row, 1, line, txt_style)
    workbook.save(filename)

def readFromXLSX(filename: str) -> JSONList:
    workbook = load_workbook(filename)
    ws = workbook.active
    cols = []
    for col in range(MAXCOL):
        header = ws.cell(row=1, column=col + 1)
        if header.value is None:
            break
        name = header.value
        if name is None:
            break
        cols.append(name)
    logg.debug("xlsx found %s cols\n\t%s", len(cols), cols)
    data: JSONList = []
    for atrow in range(MAXROWS):
        record = []
        found = 0
        for atcol in range(len(cols)):
            cell = ws.cell(row=atrow + 2, column=atcol + 1)
            value = cell.value
            # logg.debug("[%i,%si] cell.value = %s", atcol, atrow, value)
            if value is not None:
                found += 1
            if isinstance(value, str) and value == _Empty_String:
                value = ""
            record.append(value)
        if not found:
            break
        newrow = dict(zip(cols, record))
        data.append(newrow)
    return data
